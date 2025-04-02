# Methods to write Performance Tracker.

# Proposed workflow:
# - Bridge lookahead's day fraction by phase to DCCS via Event.
# - Create adjusted day fraction by phase by Event.

import openpyxl
from openpyxl.utils.cell import get_column_letter
from datetime import datetime
from settings import *


def read_DCCS(excel_file_path):
    try:
        sheet_name = 'DCCS'
        wb = openpyxl.load_workbook(filename=excel_file_path, data_only=True)
        sheet = wb[sheet_name]
        sheet.delete_rows(1, 10)
        rows_list = [row for row in sheet.iter_rows(values_only=True)]
        df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])
        df.columns = [col.date() if isinstance(col, (datetime, pd.Timestamp)) else col for col in df.columns]
        wb.close()
        return df
    except Exception as e:
        print(f"Error:", e)


def read_day_fraction(excel_file_path=TODAY_DCCS_DIR):
    try:
        sheet_name = 'Day Fraction by Phase'
        wb = openpyxl.load_workbook(filename=excel_file_path, data_only=True)
        sheet = wb[sheet_name]
        rows_list = [row for row in sheet.iter_rows(values_only=True)]
        df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])
        df.columns = [col.date() if isinstance(col, (datetime, pd.Timestamp)) else col for col in df.columns]
        wb.close()
        return df
    except Exception as e:
        print(f"Error:", e)


def first_datetime_column(df):
    return next((col for col in df.columns if pd.to_datetime(col, errors='coerce') is not pd.NaT), None)


# Load TODAY's DCCS.
df_DCCS = read_DCCS(TODAY_DCCS_DIR)
# Unpivot date columns.
df_DCCS_melted = pd.melt(df_DCCS,
                         id_vars=df_DCCS.columns[:df_DCCS.columns.get_loc(first_datetime_column(df_DCCS))],
                         var_name='Date', value_name='Quantity')
# Remove unnecessary rows and columns.
df_DCCS_melted = df_DCCS_melted.replace({pd.np.nan: 0})
df_DCCS_melted = df_DCCS_melted[df_DCCS_melted['Quantity'] != 0]
df_DCCS_melted.drop(columns=['Daily Estimate (USD)', 'Charging Mechanism', 'Total Cost (USD)', 'Total Units'],
                    inplace=True)
df_DCCS_melted.reset_index(inplace=True, drop=True)
# Generate daily line cost.
df_DCCS_melted['Daily Line Cost (USD)'] = df_DCCS_melted.apply(
    lambda row: row['Quantity'] * row['SAP Unit Price'] / (1 if row['Currency'] == "USD" else USDMYR), axis=1)

# Load TODAY's day fraction generated from lookahead.
df_day_fraction = read_day_fraction()
# Unpivot date columns.
df_day_fraction_melted = pd.melt(df_day_fraction,
                                 id_vars=df_day_fraction.columns[
                                         :df_day_fraction.columns.get_loc(first_datetime_column(df_day_fraction))],
                                 var_name='Date', value_name='Day Fraction')
# Remove unnecessary rows and columns.
df_day_fraction_melted = df_day_fraction_melted[df_day_fraction_melted['Day Fraction'] != 0]
df_day_fraction_melted.drop(columns=['Projection Start Time', 'Projection End Time', 'AFE Time', 'AFE Cost',
                                     'Actual Time', 'Days Ahead/Behind', 'Planned Depth'], inplace=True)
df_day_fraction_melted.reset_index(inplace=True, drop=True)
# Generate day fraction by Event.
df_day_fraction_melted['Day Fraction by Event'] = df_day_fraction_melted['Day Fraction'] / df_day_fraction_melted.groupby(['Well Name', 'Event', 'Date'])['Day Fraction'].transform('sum')

# Generate Phase Code for each DCCS row.
df_DCCS_expanded = pd.merge(df_DCCS.drop(columns=['Daily Estimate (USD)', 'Total Cost (USD)', 'Total Units']),
                            df_AFE[['Well Name', 'Event', 'Phase Code']],
                            left_on=['Well Name', 'Event'],
                            right_on=['Well Name', 'Event'],
                            how='left')
# Remove unnecessary datetime columns.
df_DCCS_expanded = df_DCCS_expanded.loc[:, ~pd.to_datetime(df_DCCS_expanded.columns, errors='coerce').notna()]
# Generate date and day fraction for each Phase Code.
df_DCCS_expanded = pd.merge(df_DCCS_expanded, df_day_fraction_melted,
                            left_on=['Well Name', 'Phase Code', 'Event'],
                            right_on=['Well Name', 'Phase Code', 'Event'], how='left')
# Generate daily line cost for each date.
df_DCCS_expanded = pd.merge(df_DCCS_expanded,
                            df_DCCS_melted[['Well Name', 'Event', 'OCS Number', 'Item Number', 'Description',
                                            'Date', 'Quantity', 'Daily Line Cost (USD)']],
                            on=['Well Name', 'Event', 'OCS Number', 'Item Number', 'Description', 'Date'],
                            how='left')
# Remove unnecessary rows.
df_DCCS_expanded = df_DCCS_expanded.replace({pd.np.nan: 0})
df_DCCS_expanded = df_DCCS_expanded[df_DCCS_expanded["Quantity"] != 0]
df_DCCS_expanded.reset_index(inplace=True, drop=True)
# Generate line cost for each date.  # TODO: - If cost is tagged to a Phase, check against Phase and carry full cost.
df_DCCS_expanded['Line Cost (USD)'] = df_DCCS_expanded['Daily Line Cost (USD)'] * df_DCCS_expanded['Day Fraction by Event']

# Label Actual or Projected based on date.
df_DCCS_expanded['Actual/Projected'] = df_DCCS_expanded.apply(
    lambda row: 'Actual' if row['Date'] < TODAY.date() else 'Projected', axis=1)

# Export Performance Tracker to EXCEL.
with pd.ExcelWriter(TODAY_DCCS_DIR, mode='a', engine="openpyxl") as writer:
    df_DCCS_expanded.to_excel(writer, sheet_name='DCCS Expanded', index=False, freeze_panes=(1, 0))

# Open Performance Tracker tab in EXCEL.
wb = openpyxl.load_workbook(TODAY_DCCS_DIR)
ws = wb['DCCS Expanded']

# Configure formatting.
for i, header in enumerate(list(df_DCCS_expanded.columns)):
    if header in ['Description', 'Phase']:
        ws.column_dimensions[get_column_letter(i + 1)].width = 40
    if header == 'Date':
        ws.column_dimensions[get_column_letter(i + 1)].width = 13
ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}1'

# Save workbook.
wb.save(TODAY_DCCS_DIR)
wb.close()
