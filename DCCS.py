# Methods to read, verify, and write from Excel DCCS.

# DCCS headers:
# - UID
# - Cost Group
# - Description
# - Allocation (Tariffs or OCS number)
# - MYR Cost
# - USD Cost
# - Unit
# - Total Cost (USD)
# - Total Units

# Metadata:
# - USD/MYR conversion rate

import re
from openpyxl.utils.cell import get_column_letter
from datetime import datetime
from lookahead import *
from OCS import *

# Proposed workflow:
# - Parse information from charging mechanisms.
# - Auto charging using parsed information.
# - Identify latest DCCS.
# - Use try-except to verify lookahead validity and raise errors.
# - Handle manual inputs before Today.
# TODO: - Handle consolidation especially different well from Today.
# - Generate Excel DCCS with Excel formula and intended formatting.

# Proposed verification:
# -

# Proposed data sources:
# - The latest Excel Campaign Lookahead
# - All OCS and revisions
# - All Tariffs and consolidation
# - The latest DCCS (to check for manual inputs)
# - OpenWells for rig rates and NPT information...?
# - How about aviation charges, mud charges, and POB charges...?

# Auto charging mechanism formats:
# - Date range:
# - XX unit/day from (start/end phase XX)/(YYYY/mm/dd) to (start/end phase XX)/(YYYY/mm/dd) for maximum XX occurrences
# - Well phase:
# - XX unit/day for {'Well1': [Phase Code, Phase Code, Phase Code], 'Well2':[Phase Code]} for maximum XX occurrences
# - Lump sum:
# - XX unit/day on (start/end phase XX)/(YYYY/mm/dd)
# Rules:
# - Material must be lump sum.
# - Tariff max occurrences (if applied) must be less than target well phase duration.
# - Well phase must be present on target dates.


# Parse information from charging mechanisms.
def create_instruction_dict(text, well):
    instruction_dict = {}
    text_split = text.split()
    instruction_dict['Number'] = float(text_split[0])
    instruction_dict['Recurrence'] = text_split[2]
    if text_split[2] == 'from':
        if re.findall(r'\b\d{4}/\d{2}/\d{2}\b', text_split[3]):
            instruction_dict['Start'] = pd.to_datetime(text_split[3], format='%Y/%m/%d %H:%M:%S').date()
        elif text_split[3] == 'start':  # Start of phase, to be converted to a date, and checked against well phase.
            instruction_dict['Start'] = grouped_df[(grouped_df['Well Name'] == well) & (grouped_df['Phase Code'] == int(text_split[5]))]['Projection Start Time'].iloc[0].date()
        elif text_split[3] == 'end':  # End of phase, to be converted to a date, and checked against well phase.
            instruction_dict['Start'] = grouped_df[(grouped_df['Well Name'] == well) & (grouped_df['Phase Code'] == int(text_split[5]))]['Projection End Time'].iloc[0].date()
        to_index = text_split.index('to')
        if re.findall(r'\b\d{4}/\d{2}/\d{2}\b', text_split[to_index+1]):
            instruction_dict['End'] = pd.to_datetime(text_split[to_index+1], format='%Y/%m/%d %H:%M:%S').date()
        elif text_split[to_index+1] == 'start':
            instruction_dict['End'] = grouped_df[(grouped_df['Well Name'] == well) & (grouped_df['Phase Code'] == int(text_split[to_index+3]))]['Projection Start Time'].iloc[0].date()
        elif text_split[to_index+1] == 'end':
            instruction_dict['End'] = grouped_df[(grouped_df['Well Name'] == well) & (grouped_df['Phase Code'] == int(text_split[to_index+3]))]['Projection End Time'].iloc[0].date()
        instruction_dict['Dict'] = None
    elif text_split[2] == 'for':
        instruction_dict['Start'] = None
        instruction_dict['End'] = None
        instruction_dict['Dict'] = eval(text[text.find('{'):text.find('}')+1])
    else:  # No recurrence.
        if re.findall(r'\b\d{4}/\d{2}/\d{2}\b', text_split[3]):
            instruction_dict['Start'] = pd.to_datetime(text_split[3], format='%Y/%m/%d %H:%M:%S').date()
        elif text_split[3] == 'start':
            instruction_dict['Start'] = grouped_df[(grouped_df['Well Name'] == well) & (grouped_df['Phase Code'] == int(text_split[5]))]['Projection Start Time'].iloc[0].date()
        elif text_split[3] == 'end':
            instruction_dict['Start'] = grouped_df[(grouped_df['Well Name'] == well) & (grouped_df['Phase Code'] == int(text_split[5]))]['Projection End Time'].iloc[0].date()
        instruction_dict['End'] = None
        instruction_dict['Dict'] = None
    if text_split[-1] == 'occurrences':
        instruction_dict['Occurrence'] = float(text_split[-2])
    else:
        instruction_dict['Occurrence'] = 99999
    return instruction_dict


# Auto charging using parsed information.
def auto_charge(row):
    try:
        d = create_instruction_dict(row['Charging Mechanism'], row['Well Name'])
        if d['Recurrence'] == 'from':
            # Filter by Well Event, get day fraction from date columns, multiply Number.
            filtered_df_by_event = grouped_df_short_by_event[(grouped_df_short_by_event['Well Name'] == row['Well Name']) & (grouped_df_short_by_event['Event'] == row['Event'])]
            for date in pd.date_range(start=d['Start'], end=d['End']):
                if date in well_date_range:
                    row[date.date()] = min(d['Number'] * filtered_df_by_event.iloc[0][date.date()], d['Occurrence'])
                    d['Occurrence'] -= min(d['Number'] * filtered_df_by_event.iloc[0][date.date()], d['Occurrence'])
                    if d['Occurrence'] == 0:
                        break
        elif d['Recurrence'] == 'for':
            # Filter by Well Phase, group and sum day fraction, multiply Number.
            filtered_df_by_dict = pd.DataFrame()
            for well, phase_list in d['Dict'].items():
                _temp = grouped_df_short[(grouped_df_short['Well Name'] == well) & (grouped_df_short['Phase Code'].isin(phase_list))].groupby(['Well Name', 'Event']).sum().reset_index()
                filtered_df_by_dict = pd.concat([filtered_df_by_dict, _temp], ignore_index=True)
            filtered_df_by_phase = filtered_df_by_dict[(filtered_df_by_dict['Well Name'] == row['Well Name']) & (filtered_df_by_dict['Event'] == row['Event'])]
            for date in well_date_range:
                row[date.date()] = min(d['Number'] * filtered_df_by_phase.sum().to_frame().T.iloc[0][date.date()], d['Occurrence'])
                d['Occurrence'] -= min(d['Number'] * filtered_df_by_phase.sum().to_frame().T.iloc[0][date.date()], d['Occurrence'])
                if d['Occurrence'] == 0:
                    break
        elif d['Recurrence'] == 'on':  # Lump sum.
            # If Well Event exists on the Start date, charge Number.
            filtered_df_by_event = grouped_df_short_by_event[(grouped_df_short_by_event['Well Name'] == row['Well Name']) & (grouped_df_short_by_event['Event'] == row['Event'])]
            if filtered_df_by_event.iloc[0][d['Start']]:
                row[d['Start']] = d['Number']
    except Exception as e:
        print(f"Charging error on row {row.name}: {row['Charging Mechanism']}! Error: {e}")
    return row


def read_DCCS(excel_file_path=LATEST_DCCS_DIR):
    try:
        sheet_name = 'DCCS'
        wb = openpyxl.load_workbook(filename=excel_file_path, data_only=True)
        sheet = wb[sheet_name]
        sheet.delete_rows(1, 10)
        rows_list = [row for row in sheet.iter_rows(values_only=True)]
        df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])
        df.columns = [col.date() if isinstance(col, (datetime, pd.Timestamp)) else col for col in df.columns]
        wb.close()
        return df
    except Exception as e:
        print(f"Error:", e)


# Handle manual inputs before Today.
def update_manual_inputs(df_old, df_new):
    df_one = df_old.copy(deep=True)
    df_two = df_new.copy(deep=True)
    # Get all column names before Description.
    column_list = list(df_one.columns[:df_one.columns.get_loc('Description')+1])
    # Set index as UID for both dataframes.
    df_one.set_index(column_list, inplace=True)
    df_two.set_index(column_list, inplace=True)
    # Define cut-off date for manual inputs from old dataframe.
    cutoff_date = TODAY - pd.Timedelta(days=2)  # Or custom date e.g. datetime(2024, 5, 1).
    # Convert datetime to date.
    df_one.columns = [col.date() if isinstance(col, (datetime, pd.Timestamp)) else col for col in df_one.columns]
    # Get all date columns before (not including) cut-off date.
    date_columns = [col for col in df_one.columns if pd.to_datetime(col, errors='coerce') < cutoff_date]
    # Update new dataframe with old dataframe for dates before cut-off date.
    df_two.update(df_one[date_columns])
    df_two.reset_index(inplace=True)
    return df_two


# Generate DCCS with placeholder columns.
df_DCCS = df_OCS.copy(deep=True)
df_DCCS['Daily Estimate (USD)'] = None
df_DCCS['Total Cost (USD)'] = None
df_DCCS['Total Units'] = None
df_DCCS['Vendor'] = 'Placeholder'  # Placeholder.
df_DCCS['Demand Category'] = 'Placeholder'  # Placeholder for Service vs Material.
DCCS_headers = ['File Name', 'Vendor', 'Well Name', 'Event',  # Keep Well Name to column C.
                'OCS Number', 'Item Number', 'WBS Number', 'Demand Category',  # Metadata, can be hidden.
                'Cost Group', 'Description', 'Daily Estimate (USD)',  # Required for EDM input.
                'SAP Unit Price', 'Currency', 'Unit of Measure',  # Keep together.
                'Charging Mechanism',
                'Total Cost (USD)', 'Total Units']  # Keep together.
df_DCCS = df_DCCS[DCCS_headers]
for date in well_date_range:
    df_DCCS[date.date()] = None

# Custom dataframes for auto-charging.
grouped_df_short = grouped_df.drop(
    columns=['Phase', 'Projection Start Time', 'Projection End Time', 'AFE Time', 'AFE Cost',
             'Actual Time', 'Days Ahead/Behind', 'Planned Depth'])
grouped_df_short_by_event = grouped_df_short.groupby(['Well Name', 'Event']).sum().reset_index()

# Charge DCCS as per charging mechanisms.
df_DCCS = df_DCCS.apply(lambda row: auto_charge(row), axis=1).replace({pd.np.nan: None, 0: None})

# Handle manual inputs before Today.
try:
    df_old_DCCS = read_DCCS()
    df_DCCS = update_manual_inputs(df_old_DCCS, df_DCCS)
except Exception as e:
    print("Error:", e)

# Create EXCEL formulas.
date_col_index = len(DCCS_headers)+1
sum_col_index = df_DCCS.columns.get_loc('Total Units')+1
price_col_index = df_DCCS.columns.get_loc('SAP Unit Price')+1
description_col_index = df_DCCS.columns.get_loc('Description')+1
well_col_index = df_DCCS.columns.get_loc('Well Name')+1
start_row = 10
df_DCCS['Total Units'] = df_DCCS.apply(lambda row: f'=SUM({get_column_letter(date_col_index)}{row.name + start_row + 2}:{get_column_letter(len(df_DCCS.columns))}{row.name + start_row + 2})', axis=1)
df_DCCS['Total Cost (USD)'] = df_DCCS.apply(lambda row: '={col1}{row1}*{col2}{row1}/IF({col3}{row1}="USD",1,$C$8)'.format(col1=get_column_letter(price_col_index), row1=row.name + start_row + 2, col2=get_column_letter(sum_col_index), col3=get_column_letter(price_col_index + 1)), axis=1)
df_DCCS['Daily Estimate (USD)'] = df_DCCS.apply(lambda row: '=({col1}{row1}=$C$6)*HLOOKUP($C$5,${col2}${row2}:${col3}{row1},ROW({col1}{row1})-{startrow},FALSE)*{col4}{row1}/IF({col5}{row1}="USD",1,$C$8)'.format(col1=get_column_letter(well_col_index), row1=row.name + start_row + 2, col2=get_column_letter(date_col_index), row2=start_row + 1, col3=get_column_letter(len(df_DCCS.columns)), col4=get_column_letter(price_col_index), col5=get_column_letter(price_col_index + 1), startrow=start_row), axis=1)

# Remove all empty columns.
df_DCCS = df_DCCS.loc[:, ~((df_DCCS == 0) | (df_DCCS.isna()) | (df_DCCS == '')).all()]

# Export to EXCEL.
with pd.ExcelWriter(TODAY_DCCS_DIR) as writer:
    df_DCCS.to_excel(writer, sheet_name='DCCS', index=False, startrow=start_row,
                     freeze_panes=(start_row + 1, date_col_index-1))
    grouped_df.to_excel(writer, sheet_name='Day Fraction by Phase', index=False)

# Open DCCS tab in EXCEL.
wb = openpyxl.load_workbook(TODAY_DCCS_DIR)
ws = wb['DCCS']

# Configure formatting.
ws.column_dimensions[get_column_letter(well_col_index)].width = 13
ws.column_dimensions['C'].width = 13
ws.column_dimensions[get_column_letter(description_col_index)].width = 40
ws.column_dimensions.group('E', 'H', hidden=True)
for col_idx in range(date_col_index, ws.max_column+1):
    ws.column_dimensions[get_column_letter(col_idx)].width = 13
    ws[f'{get_column_letter(col_idx)}{start_row+1}'].alignment = openpyxl.styles.Alignment(horizontal='left')
    days_before_today = (TODAY - ws[f'{get_column_letter(col_idx)}{start_row+1}'].value).days
    yellow_fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    if days_before_today > 1:
        ws[f'{get_column_letter(col_idx)}{start_row+1}'].fill = yellow_fill
    if days_before_today == 1:
        ws[f'{get_column_letter(col_idx)}{start_row+1}'].fill = green_fill
    if days_before_today == 6:
        ws.column_dimensions.group(get_column_letter(date_col_index), get_column_letter(col_idx), hidden=True)
ws.auto_filter.ref = f'A{start_row+1}:{get_column_letter(df_DCCS.shape[1])}{start_row+1}'

# Write to cells.
ws["B5"] = "Date"
ws["B6"] = "Well"
ws["B8"] = "USDMYR"
ws["B9"] = "Total well cost (USD)"
ws[f'{get_column_letter(date_col_index-1)}9'] = "Daily cost by well (USD)"
ws[f'{get_column_letter(date_col_index-1)}9'].alignment = openpyxl.styles.Alignment(horizontal='right')
ws["C5"] = TODAY.date()-pd.Timedelta(days=1)
try:
    ws["C6"] = grouped_df[grouped_df[TODAY.date()-pd.Timedelta(days=1)] != 0]['Well Name'].unique()[0]
except Exception as e:
    print("Error:", e)
    ws["C6"] = grouped_df['Well Name'].unique()[0]
ws["C8"] = USDMYR
ws["C9"] = f'=SUM({get_column_letter(date_col_index)}{start_row-1}:{get_column_letter(ws.max_column+1)}{start_row-1})'
ws["C9"].number_format = '#,##0'
for col_idx in range(date_col_index, ws.max_column+1):
    ws[f'{get_column_letter(col_idx)}{start_row-1}'] = '=SUMPRODUCT(${col1}${row1}:${col1}${row2}, 1/((--(${col2}${row1}:${col2}${row2}="USD"))*(1-$C$8)+$C$8),{col3}${row1}:{col3}${row2},--(${col4}${row1}:${col4}${row2}=$C$6))'.format(
        col1=get_column_letter(price_col_index),
        row1=start_row+2,
        row2=start_row+1 + len(df_DCCS.index),
        col2=get_column_letter(price_col_index+1), col3=get_column_letter(col_idx),
        col4=get_column_letter(well_col_index))
    ws[f'{get_column_letter(col_idx)}{start_row-1}'].number_format = '#,##0.00'

# Open Day Fraction tab in EXCEL.
ws = wb['Day Fraction by Phase']

# Configure formatting.
for i, header in enumerate(list(grouped_df.columns)):
    if header == 'Planned Depth':
        ws.freeze_panes = f'{get_column_letter(i + 2)}2'
        for col_idx in range(i + 2, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 13
            ws[f'{get_column_letter(col_idx)}1'].alignment = openpyxl.styles.Alignment(horizontal='left')
    if header == 'Phase':
        ws.column_dimensions[get_column_letter(i + 1)].width = 40
    if header in ['Projection Start Time', 'Projection End Time']:
        ws.column_dimensions[get_column_letter(i + 1)].width = 20
for cell in ws[get_column_letter(grouped_df.columns.get_loc('Days Ahead/Behind')+1)]:
    if isinstance(cell.value, (int, float)):
        cell.number_format = '0.00'
ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}1'

# Save workbook.
wb.save(TODAY_DCCS_DIR)
wb.close()
