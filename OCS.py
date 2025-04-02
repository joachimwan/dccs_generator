# Methods to read and verify Excel OCS.

# OCS headers:
# - Item Number
# - Description
# - SAP Element Number (not required)
# - Quantity
# - Unit of Measure
# - Estimated Duration
# - Currency
# - SAP Unit Price
# - Total Price
# - Remarks
# - Cost Group
# - Event (e.g. DRO/ COM/ ABA) (if not assigned, generate for each well)
# - Charging Mechanism

# Metadata:
# - Contract Title (not required)
# - Contract Number (not required)
# - SAP OA Number (not required)
# - OCS Number
# - Well Name
# - WBS Number
# - Vendor

# Charging mechanism:
# - 'XXX' unit/day from 'start date to end date' for max XXX occurrences
# - 'XXX' unit/day for [list of 'Well-Phase'] for max XXX occurrences
# - 'XXX' unit/day on 'date'
# - Rig rates: Charge 1 daily from 'start date to end date' or for [list of 'Well-Phase']
# - Tariffs (e.g. ROE, Overhead): Charge 1 daily for [list of 'Well-Phase']
# - Tariffs (e.g. RTOC, LMP, Vessels): Charge 1 daily from 'start date' to 'end date'
# - Tariffs (e.g. Insurance): Lump sum charges
# - Aviation: Charge XXX on chopper days (or Manual input)
# - Consolidation for each OCS/Tariff: Manual input
# - Daily mud cost: Manual input
# - Lump sum charges: Charge XXX on 'start Well-Phase'
# - Installed equipment: Charge XXX on 'end Well-Phase'
# - Personnel and equipment rental: Charge XXX daily for [list of 'Well-Phase'] for max XXX occurrences
# - Some are specific to Well-Phase e.g. DD or TRS, some are continuous e.g. Mud logging or SCE rentals

import openpyxl
from settings import *

# Proposed workflow:
# - For each OCS file in the OCS folder, read each OCS file.
# TODO: - Use try-except to verify OCS validity and raise errors.
# TODO: - Detect OCS revisions.
# - Generate OCS rows for tariffs.
# TODO: - Differentiate spread rate (and pseudo-spread) from lump sum.

# Proposed verification:
# - All required fields are present, e.g. OCS Number, Event.
# - Event and Well Name (if present) are correct.
# - All OCS has unique OCS Number.
# - All Item Number within the same OCS are unique.


def read_OCS(excel_file_path):  # TODO: - Add 'Vendor' and 'Demand Category' to aid allocation.
    try:
        sheet_name = 'OCS Input'
        table_name = 'OCSTable'
        wb = openpyxl.load_workbook(filename=excel_file_path, data_only=True)
        ws = wb[sheet_name]
        lookup_table = ws.tables[table_name]
        data = ws[lookup_table.ref]
        rows_list = [[col.value for col in row] for row in data]
        df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])
        df['Well Name'] = ws['B5'].value
        df['OCS Number'] = ws['B4'].value
        df['WBS Number'] = ws['B6'].value
        df['File Name'] = excel_file_path.name
        wb.close()
        return df
    except Exception as e:
        print(f"Error on {excel_file_path} :", e)


# Load all OCS into a dataframe. Sorted by filename.
df_OCS = pd.concat([read_OCS(f) for f in sorted(OCS_DIR.iterdir(), key=lambda x: x.name)], ignore_index=True)

# Generate OCS rows per Event for each Tariff (i.e. no specified Event).
for index, row in df_OCS[df_OCS['Event'].isna()].iterrows():
    for well in df_AFE['Well Name'].unique() if pd.isna(row['Well Name']) else [row['Well Name']]:
        for event in df_AFE[df_AFE['Well Name'] == well]['Event'].unique():
            _row = row.copy()
            _row['Event'] = event
            _row['Well Name'] = well
            _row['Description'] = _row['Description'] + " / " + well + " / " + event
            df_OCS = pd.concat([df_OCS, _row.to_frame().T], ignore_index=True)

# Remove original Tariff and sort OCS.
df_OCS.dropna(subset=['Event'], inplace=True)
df_OCS.sort_values(by=['File Name', 'Item Number'], inplace=True, ignore_index=True)
